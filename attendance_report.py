#!/usr/bin/env python3
"""从原始考勤文件生成月度考勤汇总表。

依赖安装:
    python3 -m pip install pandas openpyxl xlrd holidays chinese-calendar

运行方式:
    python3 attendance_report.py

如需适配其他导出格式，请优先修改“配置区”。
"""

from __future__ import annotations

import calendar
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import holidays  # type: ignore
except Exception:  # pragma: no cover
    holidays = None

try:
    import chinese_calendar as cn_calendar  # type: ignore
except Exception:  # pragma: no cover
    cn_calendar = None


# =========================
# 配置区（可维护集中配置）
# =========================

# 文件配置
DATA_DIR = "data/月度文件"
INPUT_FILE = "data/月度文件/考勤打卡记录表.xls"
LEAVE_FILE = "data/月度文件/请假记录表.xls"
ANNUAL_LEAVE_FILE = "data/月度文件/员工年假总数表.xlsx"
OUTPUT_FILE = "考勤统计结果.xlsx"

ATTENDANCE_FILE_PATTERNS = ["考勤打卡记录表.xls", "考勤打卡记录表.xlsx", "*考勤*打卡*.xls", "*考勤*打卡*.xlsx", "*考勤*.xls", "*考勤*.xlsx", "a.xls", "a.xlsx"]
LEAVE_FILE_PATTERNS = ["请假记录表.xls", "请假记录表.xlsx", "*请假*.xls", "*请假*.xlsx", "b.xls", "b.xlsx"]
ANNUAL_LEAVE_FILE_PATTERNS = ["当前员工年假总数表.xls", "当前员工年假总数表.xlsx", "员工年假总数表.xls", "员工年假总数表.xlsx", "*年假*.xls", "*年假*.xlsx", "c.xls", "c.xlsx"]
CURRENT_ANNUAL_LEAVE_FILE_PREFERRED_NAMES = [
    "当前员工年假总数表.xlsx",
    "当前员工年假总数表.xls",
    "员工年假总数表.xlsx",
    "员工年假总数表.xls",
]
MONTHLY_DIR_PATTERNS = [r"^(20\d{2})-(\d{1,2})$", r"^(20\d{2})_(\d{1,2})$", r"^(20\d{2})年(\d{1,2})月$"]

# 月份配置
# 年月统一从考勤打卡记录表自动识别；若识别不到则直接报错，避免误统计到错误月份。

# 工作日配置
# 手工覆盖建议按年份维护，后续切到 2027/2028 时直接新增对应年份即可。
WORKDAY_OVERRIDES_BY_YEAR = {
    2026: {
        "extra_workdays": {
            "2026-02-14",
            "2026-02-28",
        },
        "excluded_workdays": {
            # 2026 年 2 月春节假期（可按公司安排调整）
            "2026-02-16",
            "2026-02-17",
            "2026-02-18",
            "2026-02-19",
            "2026-02-20",
            "2026-02-23",
        },
    },
}
# 工作日来源：
# - auto: 优先 chinese_calendar（含调休），否则 holidays（仅法定假日），否则按周一到周五
# - chinese_calendar: 强制使用 chinese_calendar（未安装则报错）
# - holidays: 强制使用 holidays（未安装则报错，且不含调休，需配手工覆盖）
# - manual: 仅按周一到周五 + 手工覆盖
WORKDAY_DATA_SOURCE = "auto"

# 缺失记录默认值
MISSING_DAY_OUTPUT = ""

# 列名候选配置（用于识别布局）
EMP_ID_COLUMN_CANDIDATES = ["工号", "员工工号", "编号"]
EMP_NAME_COLUMN_CANDIDATES = ["姓名", "员工姓名"]
DATE_COLUMN_CANDIDATES = ["日期", "日", "日星期"]
AM_IN_COLUMN_CANDIDATES = ["上午上班", "上班", "签到"]
AM_OUT_COLUMN_CANDIDATES = ["上午下班", "下班", "签退"]
PM_IN_COLUMN_CANDIDATES = ["下午上班", "上班", "签到"]
PM_OUT_COLUMN_CANDIDATES = ["下午下班", "下班", "签退"]

# 针对当前考勤打卡记录表的结构识别配置
DEFAULT_BLOCK_WIDTH = 15
DEFAULT_NAME_LABEL_OFFSET_TO_BLOCK_START = 8
DEFAULT_DETAIL_HEADER_ROW = 11

# 策略配置
LATE_AND_EARLY_DISPLAY_STATUS = "迟到"
COUNT_LATE_EARLY_WHEN_MISSING = False

# 由打卡时间推断状态的阈值（可根据公司制度调整）
# 早退判定使用严格小于比较，因此 16:59:59 不算早退，只有更早才算。
LATE_THRESHOLD = "08:30:00"
EARLY_THRESHOLD = "16:59:59"
INVALID_AM_IN_AFTER = "13:00:00"
INVALID_PM_OUT_BEFORE = "11:00:00"
AFTERNOON_MISSING_EARLIEST_BEFORE = "08:31:00"
AFTERNOON_MISSING_LATEST_BEFORE = "16:29:00"
MORNING_MISSING_EARLIEST_AFTER = "09:01:00"
MORNING_MISSING_LATEST_AFTER = "16:59:00"
MORNING_MISSING_IN_THRESHOLD = "11:00:00"
AFTERNOON_MISSING_OUT_THRESHOLD = "16:00:00"


ALLOWED_DAY_STATUS = {"正常", "迟到", "早退", "未打卡", "上午未打卡", "下午未打卡"}
SUMMARY_ROWS = ["正常数", "迟到", "早退", "迟到+早退", "未打卡"]
LEAVE_TYPE_ROWS = [
    "年假",
    "年假（奖励）",
    "年假（其他）",
    "事假",
    "病假",
    "婚假",
    "产假",
    "丧假",
    "探亲假",
    "公假",
    "育儿假",
]
WEEKDAY_LABELS = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
MISSING_RELATED_STATUS = {"未打卡", "上午未打卡", "下午未打卡"}
MISSING_HIGHLIGHT_COLOR = "FFFFC7CE"  # #FFC7CE
LATE_ROW_HIGHLIGHT_COLOR = "FFFFEB9C"  # #FFEB9C
EARLY_HIGHLIGHT_COLOR = "FFEAF2F8"
LATE_EARLY_HIGHLIGHT_COLOR = "FFFFF2CC"
STATUS_FILL_COLORS = {
    "迟到": LATE_ROW_HIGHLIGHT_COLOR,
    "早退": EARLY_HIGHLIGHT_COLOR,
    "迟到+早退": LATE_EARLY_HIGHLIGHT_COLOR,
    "年假": "FFE2F0D9",
    "年假（奖励）": "FFD9EAD3",
    "年假（其他）": "FFEDE7F6",
    "事假": "FFFCE5CD",
    "病假": "FFFDE9D9",
    "婚假": "FFF4E1F1",
    "产假": "FFFCE5F0",
    "丧假": "FFEAD1DC",
    "探亲假": "FFF9EAD3",
    "公假": "FFE3F2FD",
    "育儿假": "FFE8F5E9",
}

LEAVE_APPROVED_STATUSES = {"审批中", "审批通过"}
LEAVE_HALF_AM = "am"
LEAVE_HALF_PM = "pm"
FORMAL_SHEET_NAME = "正式员工汇总"
FORMAL_TITLE_TEMPLATE = "财务公司{year}年{month}月考勤表"
ANNUAL_SHEET_NAME_TEMPLATE = "{year}合计"
MONTHLY_DETAIL_SHEET_NAME_TEMPLATE = "{year}-{month:02d}考勤明细"
MONTHLY_SUMMARY_SHEET_NAME_TEMPLATE = "{year}-{month:02d}月度统计"
ANNUAL_TITLE_TEMPLATE = "财务公司{year}年度考勤表"
MONTHLY_DETAIL_TITLE_TEMPLATE = "财务公司{year}年{month}月考勤明细表"
FORMAL_REQUIRED_NAMES = ["王秋劲", "宋国华", "闫风", "赖建阳"]
FORMAL_SEQ_COLUMN_CANDIDATES = ["序号"]
FORMAL_NAME_COLUMN_CANDIDATES = ["姓名"]
ANNUAL_LEAVE_TOTAL_COLUMN_CANDIDATES = ["年假总天数", "年假天数", "年假"]
ANNUAL_LEAVE_BALANCE_ROWS = ["年假"]
FORMAL_NON_PUBLIC_LEAVE_TOTAL_LABEL = "公假外各类休假总计"
FORMAL_SUMMARY_COLUMN_SPECS = [
    ("正常数", "正常数"),
    ("迟到", "迟到"),
    ("早退", "早退"),
    ("迟到+早退", "迟到+早退"),
    ("事假", "事假"),
    ("病假", "病假"),
    ("未打卡", "未打卡"),
    ("事假+病假+未打卡", "事+病+未打卡总计"),
    ("婚假", "婚假"),
    ("产假", "产假"),
    ("丧假", "丧假"),
    ("探亲假", "探亲假"),
    ("公假", "公假"),
    ("育儿假", "育儿假"),
    ("年假", "年假"),
    ("年假（奖励）", "年假（奖励）"),
    ("年假（其他）", "年假（其他）"),
]
FORMAL_HEADER_FILL_COLORS = {
    "identity": "FFD9E2F3",
    "normal": "FFE2F0D9",
    "attendance": "FFFFF2CC",
    "absence": "FFFCE4D6",
    "leave": "FFEDE7F6",
    "annual": "FFE2F0D9",
}

_YEAR_WORKDAY_CACHE: Dict[int, set[date]] = {}
_YEAR_WORKDAY_SOURCE_CACHE: Dict[int, str] = {}


@dataclass
class EmployeeBlock:
    sheet_name: str
    block_start: int
    block_width: int
    name_row: int
    name_label_col: int
    name_value_col: int
    id_row: int
    id_label_col: int
    id_value_col: int
    detail_header_row: int
    detail_start_row: int
    detail_end_row: int
    day_col: int
    am_in_col: int
    am_out_col: int
    pm_in_col: int
    pm_out_col: int


@dataclass
class FormalEmployeeLeaveInfo:
    seq: int
    name: str
    annual_leave_total: float


@dataclass
class MonthlySourceBundle:
    year: int
    month: int
    attendance_file: Path
    leave_file: Path
    annual_leave_file: Path


@dataclass
class MonthFolderInspection:
    folder_name: str
    year: int
    month: int
    attendance_files: List[Path]
    leave_files: List[Path]
    annual_files: List[Path]
    detail: str
    ready: bool
    has_any_data: bool


@dataclass
class MonthlyProcessedResult:
    bundle: MonthlySourceBundle
    employee_name_by_id: Dict[str, str]
    formal_employees: List[FormalEmployeeLeaveInfo]
    report_rows: List[List[object]]
    formal_summary_rows: List[List[object]]
    summary_by_emp: Dict[str, Dict[str, float]]
    workday_source: str
    leave_record_count: int


@dataclass
class ReportSummary:
    year: int
    monthly_results: List[MonthlyProcessedResult]
    formal_employee_count: int
    output_file: Path


@dataclass
class RefreshSummary:
    years: List[int]
    refreshed_months: List[Tuple[int, int]]
    output_file: Path


@dataclass
class AnnualFormalEmployeeInfo:
    seq: int
    name: str
    annual_leave_total: float
    employment_status: str


def _norm_label(value: object) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", "", text).strip()


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _parse_hms(text: str) -> time:
    return datetime.strptime(text, "%H:%M:%S").time()


def _extract_times(value: object) -> List[time]:
    text = _clean_text(value)
    if not text:
        return []
    times: List[time] = []
    for match in re.finditer(r"(\d{1,2}):(\d{2}):(\d{2})", text):
        hour, minute, second = map(int, match.groups())
        if hour > 23 or minute > 59 or second > 59:
            continue
        times.append(time(hour, minute, second))
    return times


def _time_ge(t1: time, t2: time) -> bool:
    return (t1.hour, t1.minute, t1.second) >= (t2.hour, t2.minute, t2.second)


def _time_le(t1: time, t2: time) -> bool:
    return (t1.hour, t1.minute, t1.second) <= (t2.hour, t2.minute, t2.second)


def _time_gt(t1: time, t2: time) -> bool:
    return (t1.hour, t1.minute, t1.second) > (t2.hour, t2.minute, t2.second)


def _time_lt(t1: time, t2: time) -> bool:
    return (t1.hour, t1.minute, t1.second) < (t2.hour, t2.minute, t2.second)


def _parse_day_from_cell(text: str) -> Optional[int]:
    match = re.search(r"(\d{1,2})", text)
    if not match:
        return None
    day = int(match.group(1))
    if 1 <= day <= 31:
        return day
    return None


def _excel_engine_for_path(file_path: str | Path) -> Optional[str]:
    suffix = Path(file_path).suffix.lower()
    if suffix == ".xls":
        return "xlrd"
    return None


def _extract_year_month_from_text(text: str) -> Optional[Tuple[int, int]]:
    normalized = _clean_text(text)
    patterns = [
        r"(20\d{2})[-_/\.年](\d{1,2})月?",
        r"(20\d{2})(\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, normalized)
        if not match:
            continue
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month
    return None


def read_input_file(file_path: str = INPUT_FILE) -> Dict[str, pd.DataFrame]:
    """读取考勤打卡记录表并返回所有 sheet。"""
    engine = _excel_engine_for_path(file_path)
    xls = pd.ExcelFile(file_path, engine=engine)
    return {
        sheet_name: pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            dtype=object,
            engine=engine,
        )
        for sheet_name in xls.sheet_names
    }


def read_leave_file(file_path: str = LEAVE_FILE) -> pd.DataFrame:
    """读取请假记录表，请假表默认只有一个 sheet。"""
    path = Path(file_path)
    if not path.exists():
        return pd.DataFrame()
    engine = _excel_engine_for_path(path)
    xls = pd.ExcelFile(path, engine=engine)
    if not xls.sheet_names:
        return pd.DataFrame()
    return pd.read_excel(path, sheet_name=xls.sheet_names[0], dtype=str, engine=engine).fillna("")


def read_annual_leave_file(file_path: str = ANNUAL_LEAVE_FILE) -> pd.DataFrame:
    """读取员工年假总数表，默认首行可能为空，因此不直接使用 header。"""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"年假文件不存在: {file_path}")
    engine = _excel_engine_for_path(path)
    return pd.read_excel(path, header=None, dtype=object, engine=engine).fillna("")


def _detect_target_year_month(sheet_map: Dict[str, pd.DataFrame]) -> Optional[Tuple[int, int]]:
    pattern = re.compile(r"(20\d{2})-(\d{2})-\d{2}[～~](20\d{2})-(\d{2})-\d{2}")
    for df in sheet_map.values():
        max_r = min(df.shape[0], 10)
        max_c = min(df.shape[1], 40)
        for r in range(max_r):
            for c in range(max_c):
                txt = _clean_text(df.iat[r, c])
                if not txt:
                    continue
                match = pattern.search(txt)
                if match:
                    y1, m1, _, _ = match.groups()
                    return int(y1), int(m1)
    return None


def resolve_target_year_month(sheet_map: Dict[str, pd.DataFrame]) -> Tuple[int, int]:
    detected = _detect_target_year_month(sheet_map)
    if detected:
        return detected
    raise RuntimeError("未能从考勤打卡记录表识别统计年月，请检查表头日期范围格式。")


def _find_candidate_files(root_dir: Path, patterns: Iterable[str]) -> List[Path]:
    matches: List[Path] = []
    seen = set()
    for pattern in patterns:
        for path in root_dir.rglob(pattern):
            if not path.is_file():
                continue
            resolved = path.resolve()
            if resolved in seen:
                continue
            matches.append(path)
            seen.add(resolved)
    return sorted(matches)


def _detect_year_month_from_dir_name(dir_name: str) -> Optional[Tuple[int, int]]:
    text = _clean_text(dir_name)
    for pattern in MONTHLY_DIR_PATTERNS:
        match = re.fullmatch(pattern, text)
        if not match:
            continue
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month
    return None


def _pick_single_file(candidates: List[Path], label: str, directory: Path) -> Path:
    if not candidates:
        raise FileNotFoundError(f"{directory} 中未找到 {label} 文件。")
    if len(candidates) > 1:
        raise RuntimeError(f"{directory} 中找到多个 {label} 文件，请只保留一个: {candidates}")
    return candidates[0]


def resolve_current_annual_leave_file(root_dir: Path) -> Path:
    root_candidates = [
        path
        for path in _find_candidate_files(root_dir, ANNUAL_LEAVE_FILE_PATTERNS)
        if path.parent == root_dir
    ]
    if not root_candidates:
        fallback_file = Path(ANNUAL_LEAVE_FILE)
        if fallback_file.exists() and fallback_file.parent.resolve() == root_dir.resolve():
            return fallback_file
        raise FileNotFoundError(f"未找到当前员工年假总数表，请先上传到目录: {root_dir}")

    for preferred_name in CURRENT_ANNUAL_LEAVE_FILE_PREFERRED_NAMES:
        matches = [path for path in root_candidates if path.name == preferred_name]
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            raise RuntimeError(f"{root_dir} 中找到多个当前年假文件，请只保留一个: {matches}")

    if len(root_candidates) == 1:
        return root_candidates[0]

    raise RuntimeError(f"{root_dir} 中找到多个当前年假文件，请只保留一个: {root_candidates}")


def resolve_available_annual_leave_file(root_dir: Path) -> Path:
    try:
        return resolve_current_annual_leave_file(root_dir)
    except FileNotFoundError:
        month_candidates: List[Tuple[Tuple[int, int], Path]] = []
        child_dirs = sorted(root_dir.iterdir()) if root_dir.exists() else []
        for child in child_dirs:
            if not child.is_dir():
                continue
            year_month = _detect_year_month_from_dir_name(child.name)
            annual_files = _find_candidate_files(child, ANNUAL_LEAVE_FILE_PATTERNS)
            if year_month and len(annual_files) == 1:
                month_candidates.append((year_month, annual_files[0]))
        if not month_candidates:
            raise
        return sorted(month_candidates, key=lambda item: item[0])[-1][1]


def get_current_annual_leave_summary(data_dir: str = DATA_DIR) -> Dict[str, object]:
    root_dir = Path(data_dir)
    annual_file = resolve_available_annual_leave_file(root_dir)
    annual_leave_df = read_annual_leave_file(str(annual_file))
    formal_employees = parse_formal_employee_leave_info(annual_leave_df)
    updated_at = datetime.fromtimestamp(annual_file.stat().st_mtime)
    return {
        "path": annual_file,
        "file_name": annual_file.name,
        "employee_count": len(formal_employees),
        "updated_at": updated_at,
    }


def _discover_monthly_bundles_from_subdirs(
    root_dir: Path,
    target_year: Optional[int] = None,
    relaxed: bool = False,
) -> List[MonthlySourceBundle]:
    bundles: List[MonthlySourceBundle] = []
    current_annual_leave_file: Optional[Path] = None
    month_dirs = []
    for path in sorted(root_dir.iterdir()):
        if not path.is_dir():
            continue
        attendance_files = _find_candidate_files(path, ATTENDANCE_FILE_PATTERNS)
        leave_files = _find_candidate_files(path, LEAVE_FILE_PATTERNS)
        annual_files = _find_candidate_files(path, ANNUAL_LEAVE_FILE_PATTERNS)
        if attendance_files or leave_files:
            month_dirs.append((path, attendance_files, leave_files, annual_files))

    if not month_dirs:
        return []

    seen_keys = set()
    for month_dir, attendance_files, leave_files, annual_files in month_dirs:
        dir_year_month = _detect_year_month_from_dir_name(month_dir.name)
        if target_year is not None and dir_year_month is not None and dir_year_month[0] != target_year:
            continue

        if relaxed and (len(attendance_files) != 1 or len(leave_files) != 1 or len(annual_files) > 1):
            continue

        attendance_file = _pick_single_file(attendance_files, "考勤打卡", month_dir)
        leave_file = _pick_single_file(leave_files, "请假", month_dir)
        if len(annual_files) > 1:
            raise RuntimeError(f"{month_dir} 中找到多个年假文件，请只保留一个: {annual_files}")
        if annual_files:
            annual_leave_file = annual_files[0]
        else:
            if current_annual_leave_file is None:
                current_annual_leave_file = resolve_available_annual_leave_file(root_dir)
            annual_leave_file = current_annual_leave_file

        # 以考勤文件内容识别年月；如果目录名本身也像月份，则做一致性校验。
        sheet_map = read_input_file(str(attendance_file))
        detected_year, detected_month = resolve_target_year_month(sheet_map)
        if target_year is not None and detected_year != target_year:
            continue
        if dir_year_month is not None and (detected_year, detected_month) != dir_year_month:
            if relaxed:
                continue
            raise RuntimeError(
                f"{attendance_file} 识别到的年月为 {detected_year}-{detected_month:02d}，"
                f"与目录 {month_dir.name} 不一致。"
            )
        if (detected_year, detected_month) in seen_keys:
            if relaxed:
                continue
            raise RuntimeError(f"发现重复月份目录: {month_dir}")
        seen_keys.add((detected_year, detected_month))

        bundles.append(
            MonthlySourceBundle(
                year=detected_year,
                month=detected_month,
                attendance_file=attendance_file,
                leave_file=leave_file,
                annual_leave_file=annual_leave_file,
            )
        )

    return bundles


def inspect_month_source_folders(
    data_dir: str = DATA_DIR,
    target_year: Optional[int] = None,
) -> List[MonthFolderInspection]:
    root_dir = Path(data_dir)
    if not root_dir.exists():
        return []

    rows: List[MonthFolderInspection] = []
    for month_dir in sorted(root_dir.iterdir()):
        if not month_dir.is_dir():
            continue
        year_month = _detect_year_month_from_dir_name(month_dir.name)
        if year_month is None:
            continue
        year, month = year_month
        if target_year is not None and year != target_year:
            continue

        attendance_files = _find_candidate_files(month_dir, ATTENDANCE_FILE_PATTERNS)
        leave_files = _find_candidate_files(month_dir, LEAVE_FILE_PATTERNS)
        annual_files = _find_candidate_files(month_dir, ANNUAL_LEAVE_FILE_PATTERNS)

        problems: List[str] = []
        if not attendance_files:
            problems.append("缺少考勤文件")
        elif len(attendance_files) > 1:
            problems.append("重复考勤文件")
        if not leave_files:
            problems.append("缺少请假文件")
        elif len(leave_files) > 1:
            problems.append("重复请假文件")
        if len(annual_files) > 1:
            problems.append("重复年假文件")

        has_any_data = bool(attendance_files or leave_files or annual_files or any(month_dir.iterdir()))
        rows.append(
            MonthFolderInspection(
                folder_name=month_dir.name,
                year=year,
                month=month,
                attendance_files=attendance_files,
                leave_files=leave_files,
                annual_files=annual_files,
                detail="；".join(problems),
                ready=not problems and len(attendance_files) == 1 and len(leave_files) == 1,
                has_any_data=has_any_data,
            )
        )

    return rows


def _detect_leave_year_month(leave_df: pd.DataFrame, file_path: Path) -> Optional[Tuple[int, int]]:
    detected = _extract_year_month_from_text(file_path.stem)
    if detected:
        return detected
    if leave_df.empty:
        return None

    month_counter: Dict[Tuple[int, int], int] = defaultdict(int)
    for col_name in ("开始日期", "结束日期"):
        if col_name not in leave_df.columns:
            continue
        for value in leave_df[col_name]:
            parsed = _safe_parse_date(value)
            if parsed is None:
                continue
            month_counter[(parsed.year, parsed.month)] += 1
    if not month_counter:
        return None
    return max(month_counter.items(), key=lambda item: item[1])[0]


def _choose_companion_file(
    target_year: int,
    target_month: int,
    attendance_parent: Path,
    candidate_files: List[Path],
    detected_months: Dict[Path, Optional[Tuple[int, int]]],
    fallback_file: Path,
) -> Path:
    same_month = [
        path
        for path in candidate_files
        if detected_months.get(path) == (target_year, target_month)
    ]
    if same_month:
        same_dir = [path for path in same_month if path.parent == attendance_parent]
        return sorted(same_dir or same_month)[0]

    same_dir = [path for path in candidate_files if path.parent == attendance_parent]
    if same_dir:
        return sorted(same_dir)[0]

    if fallback_file.exists():
        return fallback_file
    raise FileNotFoundError(
        f"未找到 {target_year}-{target_month:02d} 对应的配套文件，目录: {attendance_parent}"
    )


def discover_monthly_source_bundles(
    data_dir: str = DATA_DIR,
    target_year: Optional[int] = None,
    relaxed: bool = False,
) -> List[MonthlySourceBundle]:
    root_dir = Path(data_dir)
    if not root_dir.exists():
        raise FileNotFoundError(f"数据目录不存在: {data_dir}")

    monthly_dir_bundles = _discover_monthly_bundles_from_subdirs(root_dir, target_year=target_year, relaxed=relaxed)
    if monthly_dir_bundles:
        years = {bundle.year for bundle in monthly_dir_bundles}
        if len(years) > 1 and target_year is None:
            raise RuntimeError(f"当前目录包含多个年份的月目录，请按年份分开处理: {sorted(years)}")
        return monthly_dir_bundles

    attendance_candidates = _find_candidate_files(root_dir, ATTENDANCE_FILE_PATTERNS)
    leave_candidates = _find_candidate_files(root_dir, LEAVE_FILE_PATTERNS)
    annual_candidates = _find_candidate_files(root_dir, ANNUAL_LEAVE_FILE_PATTERNS)

    if not attendance_candidates:
        raise FileNotFoundError(f"未找到考勤打卡文件，请检查目录: {data_dir}")

    leave_month_map: Dict[Path, Optional[Tuple[int, int]]] = {}
    for leave_path in leave_candidates:
        leave_month_map[leave_path] = _detect_leave_year_month(read_leave_file(str(leave_path)), leave_path)

    bundles: Dict[Tuple[int, int], MonthlySourceBundle] = {}
    for attendance_path in attendance_candidates:
        sheet_map = read_input_file(str(attendance_path))
        year, month = resolve_target_year_month(sheet_map)
        if target_year is not None and year != target_year:
            continue
        key = (year, month)
        if key in bundles:
            if relaxed:
                continue
            raise RuntimeError(
                f"发现重复月份的考勤文件: {bundles[key].attendance_file} 与 {attendance_path}"
            )

        try:
            leave_path = _choose_companion_file(
                year,
                month,
                attendance_path.parent,
                leave_candidates,
                leave_month_map,
                Path(LEAVE_FILE),
            )
            annual_path = (
                _choose_companion_file(
                    year,
                    month,
                    attendance_path.parent,
                    annual_candidates,
                    {path: _extract_year_month_from_text(path.stem) for path in annual_candidates},
                    Path(ANNUAL_LEAVE_FILE),
                )
                if annual_candidates
                else resolve_available_annual_leave_file(root_dir)
            )
        except Exception:
            if relaxed:
                continue
            raise
        bundles[key] = MonthlySourceBundle(
            year=year,
            month=month,
            attendance_file=attendance_path,
            leave_file=leave_path,
            annual_leave_file=annual_path,
        )

    years = {year for year, _ in bundles}
    if len(years) > 1 and target_year is None:
        raise RuntimeError(f"当前目录包含多个年份的数据，请按年份分开处理: {sorted(years)}")

    return [bundles[key] for key in sorted(bundles)]


def normalize_columns(sheet_map: Dict[str, pd.DataFrame]) -> List[EmployeeBlock]:
    """识别考勤打卡记录表的真实布局并标准化为员工块定义。"""
    blocks: List[EmployeeBlock] = []
    seen_positions = set()

    for sheet_name, df in sheet_map.items():
        if df.empty:
            continue

        # 识别“姓名”标签列，推算每个员工块起始列
        name_positions: List[Tuple[int, int]] = []
        max_r = min(df.shape[0], 12)
        max_c = df.shape[1]
        for r in range(max_r):
            for c in range(max_c - 1):
                if _norm_label(df.iat[r, c]) in EMP_NAME_COLUMN_CANDIDATES:
                    name_val = _clean_text(df.iat[r, c + 1])
                    if name_val:
                        name_positions.append((r, c))

        for name_row, name_label_col in name_positions:
            block_start = name_label_col - DEFAULT_NAME_LABEL_OFFSET_TO_BLOCK_START
            if block_start < 0:
                continue
            pos_key = (sheet_name, block_start)
            if pos_key in seen_positions:
                continue

            id_row = name_row + 1
            if id_row >= df.shape[0]:
                continue
            if _norm_label(df.iat[id_row, name_label_col]) not in EMP_ID_COLUMN_CANDIDATES:
                continue

            emp_id = _clean_text(df.iat[id_row, name_label_col + 1])
            emp_name = _clean_text(df.iat[name_row, name_label_col + 1])
            if not emp_id or not emp_name:
                continue

            detail_header_row = DEFAULT_DETAIL_HEADER_ROW
            detail_start_row = detail_header_row + 1
            detail_end_row = min(df.shape[0] - 1, detail_start_row + 31)

            # 读取块内“上班/下班”列，适配真实 punch 列
            label_row_values = [
                _norm_label(df.iat[detail_header_row, c])
                for c in range(block_start, min(block_start + DEFAULT_BLOCK_WIDTH, df.shape[1]))
            ]
            updown_positions = [
                idx
                for idx, v in enumerate(label_row_values)
                if v in AM_IN_COLUMN_CANDIDATES or v in AM_OUT_COLUMN_CANDIDATES
            ]
            # 正常期望: [1,3,6,8]，若识别不足则回退固定结构
            if len(updown_positions) >= 4:
                am_in_rel, am_out_rel, pm_in_rel, pm_out_rel = updown_positions[:4]
            else:
                am_in_rel, am_out_rel, pm_in_rel, pm_out_rel = 1, 3, 6, 8

            block = EmployeeBlock(
                sheet_name=sheet_name,
                block_start=block_start,
                block_width=DEFAULT_BLOCK_WIDTH,
                name_row=name_row,
                name_label_col=name_label_col,
                name_value_col=name_label_col + 1,
                id_row=id_row,
                id_label_col=name_label_col,
                id_value_col=name_label_col + 1,
                detail_header_row=detail_header_row,
                detail_start_row=detail_start_row,
                detail_end_row=detail_end_row,
                day_col=block_start,
                am_in_col=block_start + am_in_rel,
                am_out_col=block_start + am_out_rel,
                pm_in_col=block_start + pm_in_rel,
                pm_out_col=block_start + pm_out_rel,
            )
            blocks.append(block)
            seen_positions.add(pos_key)

    return sorted(blocks, key=lambda b: (b.sheet_name, b.block_start))


def _safe_parse_iso_date(value: str) -> Optional[date]:
    text = _clean_text(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _safe_parse_date(value: object) -> Optional[date]:
    text = _clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _safe_parse_time(value: object) -> Optional[time]:
    text = _clean_text(value)
    if not text:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None


def _iter_year_dates(year: int) -> List[date]:
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    return [start + timedelta(days=i) for i in range((end - start).days + 1)]


def _get_year_workday_overrides(year: int) -> Tuple[set[date], set[date]]:
    override_cfg = WORKDAY_OVERRIDES_BY_YEAR.get(year, {})
    extra_raw = override_cfg.get("extra_workdays", set())
    excluded_raw = override_cfg.get("excluded_workdays", set())
    extra = {
        d
        for d in (_safe_parse_iso_date(v) for v in extra_raw)
        if d is not None and d.year == year
    }
    excluded = {
        d
        for d in (_safe_parse_iso_date(v) for v in excluded_raw)
        if d is not None and d.year == year
    }
    return extra, excluded


def _build_year_workday_set(year: int) -> Tuple[set[date], str]:
    """先构建全年工作日集合，再供按月过滤使用。"""
    source = WORKDAY_DATA_SOURCE.lower().strip()
    year_days = _iter_year_dates(year)

    if source == "auto":
        if cn_calendar is not None:
            base_source = "chinese_calendar"
            workday_set = {d for d in year_days if cn_calendar.is_workday(d)}
        elif holidays is not None:
            base_source = "holidays"
            cn_holidays = holidays.country_holidays("CN", years=[year])
            workday_set = {d for d in year_days if d.weekday() < 5 and d not in cn_holidays}
        else:
            base_source = "manual_weekday"
            workday_set = {d for d in year_days if d.weekday() < 5}
    elif source == "chinese_calendar":
        if cn_calendar is None:
            raise RuntimeError("WORKDAY_DATA_SOURCE=chinese_calendar 但未安装 chinese-calendar")
        base_source = "chinese_calendar"
        workday_set = {d for d in year_days if cn_calendar.is_workday(d)}
    elif source == "holidays":
        if holidays is None:
            raise RuntimeError("WORKDAY_DATA_SOURCE=holidays 但未安装 holidays")
        base_source = "holidays"
        cn_holidays = holidays.country_holidays("CN", years=[year])
        workday_set = {d for d in year_days if d.weekday() < 5 and d not in cn_holidays}
    elif source == "manual":
        base_source = "manual_weekday"
        workday_set = {d for d in year_days if d.weekday() < 5}
    else:
        raise ValueError(f"未知 WORKDAY_DATA_SOURCE: {WORKDAY_DATA_SOURCE}")

    # 手工覆盖（最终优先级最高）
    extra, excluded = _get_year_workday_overrides(year)
    workday_set = (workday_set - excluded) | extra
    return workday_set, base_source


def get_year_workday_set(year: int) -> Tuple[set[date], str]:
    if year not in _YEAR_WORKDAY_CACHE:
        workday_set, source = _build_year_workday_set(year)
        _YEAR_WORKDAY_CACHE[year] = workday_set
        if WORKDAY_DATA_SOURCE.lower().strip() == "auto":
            source = f"{source}(auto)"
        _YEAR_WORKDAY_SOURCE_CACHE[year] = source
    return _YEAR_WORKDAY_CACHE[year], _YEAR_WORKDAY_SOURCE_CACHE[year]


def is_workday(current_date: date, year_workday_set: set[date]) -> bool:
    return current_date in year_workday_set


def classify_day_status(
    am_in_times: List[time],
    am_out_times: List[time],
    pm_in_times: List[time],
    pm_out_times: List[time],
) -> Dict[str, object]:
    """根据四段打卡归并为每日状态，并返回统计辅助字段。"""
    late_threshold = _parse_hms(LATE_THRESHOLD)
    early_threshold = _parse_hms(EARLY_THRESHOLD)
    invalid_am_in_after = _parse_hms(INVALID_AM_IN_AFTER)
    invalid_pm_out_before = _parse_hms(INVALID_PM_OUT_BEFORE)
    afternoon_missing_earliest_before = _parse_hms(AFTERNOON_MISSING_EARLIEST_BEFORE)
    afternoon_missing_latest_before = _parse_hms(AFTERNOON_MISSING_LATEST_BEFORE)
    morning_missing_earliest_after = _parse_hms(MORNING_MISSING_EARLIEST_AFTER)
    morning_missing_latest_after = _parse_hms(MORNING_MISSING_LATEST_AFTER)
    morning_missing_threshold = _parse_hms(MORNING_MISSING_IN_THRESHOLD)
    afternoon_missing_threshold = _parse_hms(AFTERNOON_MISSING_OUT_THRESHOLD)

    all_times = sorted(am_in_times + am_out_times + pm_in_times + pm_out_times)
    use_all_extremes = len(all_times) > 2

    # 新规则：当日有效打卡记录超过2条时，按全量最早/最晚作为上下班时间
    if use_all_extremes:
        in_times = [all_times[0]]
        out_times = [all_times[-1]]
    else:
        in_times = sorted(am_in_times + pm_in_times)
        out_times = sorted(am_out_times + pm_out_times)

    earliest_in = in_times[0] if in_times else None
    latest_out = out_times[-1] if out_times else None
    valid_punches = sorted(in_times + out_times)
    has_missing = False
    late = False
    early = False

    if not valid_punches:
        status = "未打卡"
        has_missing = True
    elif use_all_extremes:
        if earliest_in is not None and _time_gt(earliest_in, invalid_am_in_after):
            status = "未打卡"
            has_missing = True
        elif latest_out is not None and _time_lt(latest_out, invalid_pm_out_before):
            status = "未打卡"
            has_missing = True
        elif (
            earliest_in is not None
            and latest_out is not None
            and _time_lt(earliest_in, afternoon_missing_earliest_before)
            and _time_lt(latest_out, afternoon_missing_latest_before)
        ):
            status = "下午未打卡"
            has_missing = True
        elif (
            earliest_in is not None
            and latest_out is not None
            and _time_gt(earliest_in, morning_missing_earliest_after)
            and _time_gt(latest_out, morning_missing_latest_after)
        ):
            status = "上午未打卡"
            has_missing = True
        else:
            late = earliest_in is not None and _time_gt(earliest_in, late_threshold)
            early = latest_out is not None and _time_lt(latest_out, early_threshold)
            if late and early:
                status = LATE_AND_EARLY_DISPLAY_STATUS
            elif late:
                status = "迟到"
            elif early:
                status = "早退"
            else:
                status = "正常"
    elif not in_times or not out_times:
        status = "未打卡"
        has_missing = True
    elif earliest_in is not None and _time_gt(earliest_in, invalid_am_in_after):
        status = "未打卡"
        has_missing = True
    elif latest_out is not None and _time_lt(latest_out, invalid_pm_out_before):
        status = "未打卡"
        has_missing = True
    elif (
        earliest_in is not None
        and latest_out is not None
        and _time_lt(earliest_in, afternoon_missing_earliest_before)
        and _time_lt(latest_out, afternoon_missing_latest_before)
    ):
        status = "下午未打卡"
        has_missing = True
    elif (
        earliest_in is not None
        and latest_out is not None
        and _time_gt(earliest_in, morning_missing_earliest_after)
        and _time_gt(latest_out, morning_missing_latest_after)
    ):
        status = "上午未打卡"
        has_missing = True
    elif pm_in_times and not pm_out_times:
        status = "下午未打卡"
        has_missing = True
    elif not am_in_times and (am_out_times or pm_in_times or pm_out_times):
        latest_morning_out = am_out_times[-1] if am_out_times else None
        if (latest_morning_out is not None and _time_ge(latest_morning_out, morning_missing_threshold)) or pm_in_times:
            status = "上午未打卡"
        else:
            status = "未打卡"
        has_missing = True
    elif am_in_times and not am_out_times:
        status = "未打卡"
        has_missing = True
    else:
        if earliest_in is not None and _time_ge(earliest_in, morning_missing_threshold):
            status = "上午未打卡"
            has_missing = True
        elif latest_out is not None and _time_le(latest_out, afternoon_missing_threshold):
            status = "下午未打卡"
            has_missing = True
        else:
            late = earliest_in is not None and _time_gt(earliest_in, late_threshold)
            early = latest_out is not None and _time_lt(latest_out, early_threshold)
            if late and early:
                status = LATE_AND_EARLY_DISPLAY_STATUS
            elif late:
                status = "迟到"
            elif early:
                status = "早退"
            else:
                status = "正常"

    if has_missing and not COUNT_LATE_EARLY_WHEN_MISSING:
        late = False
        early = False

    if status not in ALLOWED_DAY_STATUS:
        status = "未打卡"
        has_missing = True
        late = False
        early = False

    return {
        "status": status,
        "late": int(late),
        "early": int(early),
        "late_early": int(late and early),
        "missing_score": 1.0
        if status == "未打卡"
        else 0.5
        if status in {"上午未打卡", "下午未打卡"}
        else 0.0,
        "normal": int(status == "正常"),
    }


def _severity_key(day_record: Dict[str, object]) -> Tuple[float, int, int, int, int]:
    status = str(day_record["status"])
    severity = {
        "未打卡": 5.0,
        "上午未打卡": 4.0,
        "下午未打卡": 4.0,
        "迟到": 3.0,
        "早退": 3.0,
        "正常": 2.0,
    }.get(status, 0.0)
    return (
        severity,
        int(day_record["late_early"]),
        int(day_record["late"]),
        int(day_record["early"]),
        int(day_record["normal"]),
    )


def parse_attendance_records(
    sheet_map: Dict[str, pd.DataFrame],
    employee_blocks: Iterable[EmployeeBlock],
    year: int,
    month: int,
) -> Tuple[Dict[str, str], Dict[str, Dict[date, Dict[str, object]]]]:
    """解析每个员工每天的四段打卡状态。"""
    employee_name_by_id: Dict[str, str] = {}
    records: Dict[str, Dict[date, Dict[str, object]]] = defaultdict(dict)

    _, days_in_month = calendar.monthrange(year, month)

    for block in employee_blocks:
        df = sheet_map[block.sheet_name]
        emp_id = _clean_text(df.iat[block.id_row, block.id_value_col])
        emp_name = _clean_text(df.iat[block.name_row, block.name_value_col])
        if not emp_id:
            continue
        if not emp_name:
            emp_name = f"员工{emp_id}"
        employee_name_by_id[emp_id] = emp_name

        for row in range(block.detail_start_row, block.detail_end_row + 1):
            if row >= df.shape[0]:
                break
            day_text = _clean_text(df.iat[row, block.day_col])
            if not day_text:
                continue
            day = _parse_day_from_cell(day_text)
            if day is None or day > days_in_month:
                continue
            current_date = date(year, month, day)

            day_result = classify_day_status(
                _extract_times(df.iat[row, block.am_in_col] if block.am_in_col < df.shape[1] else None),
                _extract_times(df.iat[row, block.am_out_col] if block.am_out_col < df.shape[1] else None),
                _extract_times(df.iat[row, block.pm_in_col] if block.pm_in_col < df.shape[1] else None),
                _extract_times(df.iat[row, block.pm_out_col] if block.pm_out_col < df.shape[1] else None),
            )

            old = records[emp_id].get(current_date)
            if old is None or _severity_key(day_result) > _severity_key(old):
                records[emp_id][current_date] = day_result

    return employee_name_by_id, records


def _format_stat_value(value: float) -> object:
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return round(value, 1)


def _safe_float(value: object, default: float = 0.0) -> float:
    text = _clean_text(value)
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        return float(match.group(0)) if match else default


def _find_first_matching_column(headers: List[str], candidates: Iterable[str]) -> Optional[int]:
    for idx, header in enumerate(headers):
        if header in candidates:
            return idx
    return None


def parse_formal_employee_leave_info(
    annual_leave_df: pd.DataFrame,
) -> List[FormalEmployeeLeaveInfo]:
    """解析员工年假总数表中的正式员工及年假总天数。"""
    header_row = None
    max_check_rows = min(len(annual_leave_df.index), 10)
    for r_idx in range(max_check_rows):
        row_headers = [_norm_label(v) for v in annual_leave_df.iloc[r_idx].tolist()]
        if (
            _find_first_matching_column(row_headers, FORMAL_NAME_COLUMN_CANDIDATES) is not None
            and _find_first_matching_column(row_headers, ANNUAL_LEAVE_TOTAL_COLUMN_CANDIDATES) is not None
        ):
            header_row = r_idx
            break
    if header_row is None:
        raise RuntimeError("未识别到员工年假总数表表头，请检查姓名/年假总天数列。")

    headers = [_norm_label(v) for v in annual_leave_df.iloc[header_row].tolist()]
    seq_col = _find_first_matching_column(headers, FORMAL_SEQ_COLUMN_CANDIDATES)
    name_col = _find_first_matching_column(headers, FORMAL_NAME_COLUMN_CANDIDATES)
    total_col = _find_first_matching_column(headers, ANNUAL_LEAVE_TOTAL_COLUMN_CANDIDATES)
    if name_col is None or total_col is None:
        raise RuntimeError("员工年假总数表缺少必要列：姓名 或 年假总天数。")

    rows: List[FormalEmployeeLeaveInfo] = []
    fallback_seq = 1
    for r_idx in range(header_row + 1, len(annual_leave_df.index)):
        name = _clean_text(annual_leave_df.iat[r_idx, name_col])
        if not name:
            continue
        seq_text = _clean_text(annual_leave_df.iat[r_idx, seq_col]) if seq_col is not None else ""
        seq = int(_safe_float(seq_text, fallback_seq)) if seq_text else fallback_seq
        annual_leave_total = _safe_float(annual_leave_df.iat[r_idx, total_col], 0.0)
        rows.append(FormalEmployeeLeaveInfo(seq=seq, name=name, annual_leave_total=annual_leave_total))
        fallback_seq = max(fallback_seq + 1, seq + 1)

    existing_names = {row.name for row in rows}
    for leader_name in FORMAL_REQUIRED_NAMES:
        if leader_name not in existing_names:
            rows.append(
                FormalEmployeeLeaveInfo(
                    seq=fallback_seq,
                    name=leader_name,
                    annual_leave_total=0.0,
                )
            )
            fallback_seq += 1

    return sorted(rows, key=lambda item: (item.seq, item.name))


def _normalize_leave_type(raw_leave_type: str) -> Optional[str]:
    leave_type = _clean_text(raw_leave_type)
    if not leave_type:
        return None
    alias = {
        "奖励假": "年假（奖励）",
        "年假(奖励)": "年假（奖励）",
    }
    leave_type = alias.get(leave_type, leave_type)
    return leave_type if leave_type in LEAVE_TYPE_ROWS else None


def _is_nonzero_number_text(text: str) -> bool:
    try:
        return abs(float(text)) > 1e-9
    except ValueError:
        return False


def _extract_fill_key(value_text: str) -> Optional[str]:
    text = _clean_text(value_text)
    if not text or text == "正常":
        return None
    if "未打卡" in text:
        return "未打卡"
    if text == "迟到" or text == "早退" or text == "迟到+早退":
        return text

    first_part = text.split("，", 1)[0].strip()
    first_part = re.sub(r"^(上午|下午)", "", first_part)
    ordered_leave_types = sorted(LEAVE_TYPE_ROWS, key=len, reverse=True)
    for leave_type in ordered_leave_types:
        if leave_type in first_part:
            return leave_type

    for leave_type in ordered_leave_types:
        if leave_type in text:
            return leave_type
    return None


def _strip_unit_suffix(text: str) -> str:
    return re.sub(r"（[^）]*）$", "", _clean_text(text))


def _wrap_header_text(text: object, chunk_size: int = 5) -> str:
    raw = _clean_text(text)
    special_wrap_map = {
        "年假（奖励）": 2,
        "年假（其他）": 2,
    }
    if raw in special_wrap_map and len(raw) > special_wrap_map[raw]:
        split_at = special_wrap_map[raw]
        return raw[:split_at] + "\n" + raw[split_at:]
    if len(raw) <= chunk_size:
        return raw
    return "\n".join(raw[i : i + chunk_size] for i in range(0, len(raw), chunk_size))


def get_annual_sheet_name(year: int) -> str:
    return ANNUAL_SHEET_NAME_TEMPLATE.format(year=year)


def get_monthly_detail_sheet_name(year: int, month: int) -> str:
    return MONTHLY_DETAIL_SHEET_NAME_TEMPLATE.format(year=year, month=month)


def get_monthly_summary_sheet_name(year: int, month: int) -> str:
    return MONTHLY_SUMMARY_SHEET_NAME_TEMPLATE.format(year=year, month=month)


def parse_monthly_detail_sheet_name(sheet_name: str) -> Optional[Tuple[int, int]]:
    match = re.fullmatch(r"(\d{4})-(\d{2})考勤明细", _clean_text(sheet_name))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def parse_monthly_summary_sheet_name(sheet_name: str) -> Optional[Tuple[int, int]]:
    match = re.fullmatch(r"(\d{4})-(\d{2})月度统计", _clean_text(sheet_name))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def _parse_monthly_summary_sheet_name(sheet_name: str) -> Optional[Tuple[int, int]]:
    match = re.fullmatch(r"(\d{4})-(\d{2})月度统计", _clean_text(sheet_name))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def _parse_monthly_detail_sheet_name(sheet_name: str) -> Optional[Tuple[int, int]]:
    match = re.fullmatch(r"(\d{4})-(\d{2})考勤明细", _clean_text(sheet_name))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def _overlaps_half(
    seg_start: datetime,
    seg_end: datetime,
    half_start: datetime,
    half_end: datetime,
) -> bool:
    return seg_start < half_end and seg_end > half_start


def parse_leave_records(
    leave_df: pd.DataFrame,
    employee_name_by_id: Dict[str, str],
    year: int,
    month: int,
) -> Dict[str, Dict[date, Dict[str, List[str]]]]:
    """将请假单展开到员工-日期-半天粒度。"""
    if leave_df.empty:
        return {}

    name_to_ids: Dict[str, List[str]] = defaultdict(list)
    for emp_id, name in employee_name_by_id.items():
        name_to_ids[name].append(emp_id)

    leave_col = next((c for c in leave_df.columns if "请假类型" in str(c)), None)
    if leave_col is None:
        raise RuntimeError("请假记录表未找到请假类型列")

    leave_map: Dict[str, Dict[date, Dict[str, List[str]]]] = defaultdict(
        lambda: defaultdict(lambda: {LEAVE_HALF_AM: [], LEAVE_HALF_PM: []})
    )

    target_month_start = date(year, month, 1)
    _, days_in_month = calendar.monthrange(year, month)
    target_month_end = date(year, month, days_in_month)

    for _, row in leave_df.iterrows():
        status = _clean_text(row.get("状态"))
        if status not in LEAVE_APPROVED_STATUSES:
            continue

        applicant_name = _clean_text(row.get("申请人"))
        if not applicant_name or applicant_name not in name_to_ids:
            continue

        start_date = _safe_parse_date(row.get("开始日期"))
        end_date = _safe_parse_date(row.get("结束日期"))
        start_time = _safe_parse_time(row.get("开始时间"))
        end_time = _safe_parse_time(row.get("结束时间"))
        if start_date is None or end_date is None or start_time is None or end_time is None:
            continue

        if end_date < target_month_start or start_date > target_month_end:
            continue

        leave_type = _normalize_leave_type(_clean_text(row.get(leave_col)))
        if leave_type is None:
            continue
        seg_start = datetime.combine(start_date, start_time)
        seg_end = datetime.combine(end_date, end_time)
        if seg_end <= seg_start:
            continue

        current_day = max(start_date, target_month_start)
        last_day = min(end_date, target_month_end)
        while current_day <= last_day:
            am_start = datetime.combine(current_day, time(8, 30))
            am_end = datetime.combine(current_day, time(12, 0))
            pm_start = datetime.combine(current_day, time(12, 0))
            pm_end = datetime.combine(current_day, time(17, 0))

            for emp_id in name_to_ids[applicant_name]:
                if _overlaps_half(seg_start, seg_end, am_start, am_end):
                    if leave_type not in leave_map[emp_id][current_day][LEAVE_HALF_AM]:
                        leave_map[emp_id][current_day][LEAVE_HALF_AM].append(leave_type)
                if _overlaps_half(seg_start, seg_end, pm_start, pm_end):
                    if leave_type not in leave_map[emp_id][current_day][LEAVE_HALF_PM]:
                        leave_map[emp_id][current_day][LEAVE_HALF_PM].append(leave_type)
            current_day += timedelta(days=1)

    return leave_map


def apply_leave_to_attendance(
    records: Dict[str, Dict[date, Dict[str, object]]],
    leave_map: Dict[str, Dict[date, Dict[str, List[str]]]],
) -> Dict[str, Dict[date, Dict[str, object]]]:
    """用请假半天覆盖缺卡半天，并生成请假统计。"""
    merged: Dict[str, Dict[date, Dict[str, object]]] = defaultdict(dict)

    all_emp_ids = set(records) | set(leave_map)
    for emp_id in all_emp_ids:
        all_dates = set(records.get(emp_id, {})) | set(leave_map.get(emp_id, {}))
        for current_date in all_dates:
            record = dict(records.get(emp_id, {}).get(current_date, {}))
            leave_halves = leave_map.get(emp_id, {}).get(
                current_date,
                {LEAVE_HALF_AM: [], LEAVE_HALF_PM: []},
            )

            if not record:
                base_status = "未打卡"
                record = {
                    "status": base_status,
                    "late": 0,
                    "early": 0,
                    "late_early": 0,
                    "missing_score": 1.0,
                    "normal": 0,
                }
            else:
                base_status = str(record.get("status", "未打卡"))

            # 正常出勤优先于请假
            if base_status == "正常":
                record["leave_counts"] = {leave_type: 0.0 for leave_type in LEAVE_TYPE_ROWS}
                record["display_status"] = base_status
                merged[emp_id][current_date] = record
                continue

            half_state = {
                LEAVE_HALF_AM: "work",
                LEAVE_HALF_PM: "work",
            }
            if base_status == "未打卡":
                half_state = {LEAVE_HALF_AM: "missing", LEAVE_HALF_PM: "missing"}
            elif base_status == "上午未打卡":
                half_state = {LEAVE_HALF_AM: "missing", LEAVE_HALF_PM: "work"}
            elif base_status == "下午未打卡":
                half_state = {LEAVE_HALF_AM: "work", LEAVE_HALF_PM: "missing"}

            leave_counts = {leave_type: 0.0 for leave_type in LEAVE_TYPE_ROWS}
            leave_display = {LEAVE_HALF_AM: "", LEAVE_HALF_PM: ""}

            for half in (LEAVE_HALF_AM, LEAVE_HALF_PM):
                leave_types = leave_halves.get(half, [])
                if half_state[half] == "missing" and leave_types:
                    for leave_type in leave_types:
                        leave_counts[_normalize_leave_type(leave_type)] += 0.5
                    if len(leave_types) == 1:
                        leave_display[half] = leave_types[0]
                    else:
                        leave_display[half] = "/".join(leave_types)
                    half_state[half] = "leave"

            missing_halves = sum(1 for half in half_state.values() if half == "missing")
            record["missing_score"] = missing_halves * 0.5
            if missing_halves > 0:
                record["late"] = 0
                record["early"] = 0
                record["late_early"] = 0
                record["normal"] = 0

            am_desc = ""
            pm_desc = ""
            if half_state[LEAVE_HALF_AM] == "leave":
                am_desc = f"上午{leave_display[LEAVE_HALF_AM]}"
            elif half_state[LEAVE_HALF_AM] == "missing":
                am_desc = "上午未打卡"

            if half_state[LEAVE_HALF_PM] == "leave":
                pm_desc = f"下午{leave_display[LEAVE_HALF_PM]}"
            elif half_state[LEAVE_HALF_PM] == "missing":
                pm_desc = "下午未打卡"

            if (
                half_state[LEAVE_HALF_AM] == "missing"
                and half_state[LEAVE_HALF_PM] == "missing"
            ):
                display_status = "未打卡"
            elif (
                half_state[LEAVE_HALF_AM] == "leave"
                and half_state[LEAVE_HALF_PM] == "leave"
                and leave_display[LEAVE_HALF_AM] == leave_display[LEAVE_HALF_PM]
            ):
                display_status = leave_display[LEAVE_HALF_AM]
            elif am_desc and pm_desc:
                display_status = f"{am_desc}，{pm_desc}"
            elif am_desc:
                display_status = am_desc
            elif pm_desc:
                display_status = pm_desc
            else:
                display_status = base_status

            if missing_halves == 2:
                record["status"] = "未打卡"
            elif missing_halves == 1 and half_state[LEAVE_HALF_AM] == "missing":
                record["status"] = "上午未打卡"
            elif missing_halves == 1 and half_state[LEAVE_HALF_PM] == "missing":
                record["status"] = "下午未打卡"
            elif leave_counts and sum(leave_counts.values()) > 0:
                record["status"] = base_status if base_status in {"迟到", "早退", "正常"} else "正常"
                record["late"] = 0 if "上午" in display_status and "未打卡" not in display_status else record["late"]
                record["early"] = 0 if "下午" in display_status and "未打卡" not in display_status else record["early"]
                record["late_early"] = int(record["late"] and record["early"])

            record["leave_counts"] = leave_counts
            record["display_status"] = display_status
            merged[emp_id][current_date] = record

    return merged


def _new_summary_bucket() -> Dict[str, float]:
    return {
        "迟到": 0.0,
        "早退": 0.0,
        "迟到+早退": 0.0,
        "未打卡": 0.0,
        "正常数": 0.0,
        **{leave_type: 0.0 for leave_type in LEAVE_TYPE_ROWS},
    }


def build_summary_by_employee(
    employee_ids: List[str],
    records: Dict[str, Dict[date, Dict[str, object]]],
    all_days: List[date],
    workday_set: set[date],
) -> Dict[str, Dict[str, float]]:
    summary_by_emp = {emp_id: _new_summary_bucket() for emp_id in employee_ids}
    for emp_id in employee_ids:
        for current_date in all_days:
            if current_date not in workday_set:
                continue
            day_record = records.get(emp_id, {}).get(current_date)
            if not day_record:
                continue
            summary_by_emp[emp_id]["迟到"] += float(day_record["late"])
            summary_by_emp[emp_id]["早退"] += float(day_record["early"])
            summary_by_emp[emp_id]["迟到+早退"] += float(day_record["late_early"])
            summary_by_emp[emp_id]["未打卡"] += float(day_record["missing_score"])
            summary_by_emp[emp_id]["正常数"] += float(day_record["normal"])
            for leave_type, leave_value in day_record.get("leave_counts", {}).items():
                summary_by_emp[emp_id][leave_type] += float(leave_value)
    return summary_by_emp


def build_formal_summary_rows(
    formal_employees: List[FormalEmployeeLeaveInfo],
    employee_name_by_id: Dict[str, str],
    summary_by_emp: Dict[str, Dict[str, float]],
    *,
    include_employment_status: bool = False,
    employment_status_by_name: Optional[Dict[str, str]] = None,
) -> List[List[object]]:
    columns = ["序号", "工号", "姓名"]
    if include_employment_status:
        columns.append("状态")
    columns += [label for _, label in FORMAL_SUMMARY_COLUMN_SPECS] + [
        "年假总天数",
        "结余年假总天数",
        FORMAL_NON_PUBLIC_LEAVE_TOTAL_LABEL,
    ]
    rows: List[List[object]] = [columns]

    name_to_emp_id = {name: emp_id for emp_id, name in employee_name_by_id.items()}
    leader_order = {name: idx for idx, name in enumerate(FORMAL_REQUIRED_NAMES)}
    enriched_rows = []
    for formal_employee in formal_employees:
        emp_id = name_to_emp_id.get(formal_employee.name, "")
        enriched_rows.append((formal_employee, emp_id))

    def _formal_sort_key(item: Tuple[FormalEmployeeLeaveInfo, str]) -> Tuple[int, object, int, str]:
        formal_employee, emp_id = item
        if formal_employee.name in leader_order:
            return (0, leader_order[formal_employee.name], formal_employee.seq, formal_employee.name)
        if emp_id.isdigit():
            return (1, int(emp_id), formal_employee.seq, formal_employee.name)
        if emp_id:
            return (1, emp_id, formal_employee.seq, formal_employee.name)
        return (2, formal_employee.seq, formal_employee.seq, formal_employee.name)

    for row_idx, (formal_employee, emp_id) in enumerate(sorted(enriched_rows, key=_formal_sort_key), start=1):
        summary = summary_by_emp.get(emp_id, _new_summary_bucket())
        annual_leave_used = sum(float(summary.get(label, 0.0)) for label in ANNUAL_LEAVE_BALANCE_ROWS)
        remaining_annual_leave = formal_employee.annual_leave_total - annual_leave_used
        non_public_leave_total = sum(
            float(summary.get(leave_type, 0.0))
            for leave_type in LEAVE_TYPE_ROWS
            if leave_type != "公假"
        )

        row = [
            row_idx,
            emp_id,
            formal_employee.name,
        ]
        if include_employment_status:
            row.append((employment_status_by_name or {}).get(formal_employee.name, "在职"))
        for key, _ in FORMAL_SUMMARY_COLUMN_SPECS:
            if key == "事假+病假+未打卡":
                value = (
                    float(summary.get("事假", 0.0))
                    + float(summary.get("病假", 0.0))
                    + float(summary.get("未打卡", 0.0))
                )
            else:
                value = float(summary.get(key, 0.0))
            row.append(_format_stat_value(value))
        row.append(_format_stat_value(formal_employee.annual_leave_total))
        row.append(_format_stat_value(remaining_annual_leave))
        row.append(_format_stat_value(non_public_leave_total))
        rows.append(row)

    return rows


def build_monthly_formal_summary_rows(
    formal_employees: List[FormalEmployeeLeaveInfo],
    employee_name_by_id: Dict[str, str],
    summary_by_emp: Dict[str, Dict[str, float]],
) -> List[List[object]]:
    return build_formal_summary_rows(
        formal_employees,
        employee_name_by_id,
        summary_by_emp,
        include_employment_status=True,
        employment_status_by_name={employee.name: "在职" for employee in formal_employees},
    )


def build_annual_summary_rows(
    formal_employees: List[AnnualFormalEmployeeInfo],
    employee_name_by_id: Dict[str, str],
    annual_summary_by_emp: Dict[str, Dict[str, float]],
) -> List[List[object]]:
    employment_status_by_name = {employee.name: employee.employment_status for employee in formal_employees}
    base_rows = [
        FormalEmployeeLeaveInfo(
            seq=employee.seq,
            name=employee.name,
            annual_leave_total=employee.annual_leave_total,
        )
        for employee in formal_employees
    ]
    return build_formal_summary_rows(
        base_rows,
        employee_name_by_id,
        annual_summary_by_emp,
        include_employment_status=True,
        employment_status_by_name=employment_status_by_name,
    )


def merge_annual_formal_employees(
    monthly_results: List[MonthlyProcessedResult],
) -> List[AnnualFormalEmployeeInfo]:
    merged_by_name: Dict[str, AnnualFormalEmployeeInfo] = {}
    if not monthly_results:
        return []

    last_month_names = {employee.name for employee in monthly_results[-1].formal_employees}
    for monthly_result in monthly_results:
        for employee in monthly_result.formal_employees:
            existing = merged_by_name.get(employee.name)
            if existing is None:
                merged_by_name[employee.name] = AnnualFormalEmployeeInfo(
                    seq=employee.seq,
                    name=employee.name,
                    annual_leave_total=employee.annual_leave_total,
                    employment_status="在职" if employee.name in last_month_names else "离职",
                )
                continue

            # 保留更靠前的序号，年假总数采用最后一次出现的值。
            existing.seq = min(existing.seq, employee.seq)
            existing.annual_leave_total = employee.annual_leave_total
            existing.employment_status = "在职" if employee.name in last_month_names else "离职"

    leader_order = {name: idx for idx, name in enumerate(FORMAL_REQUIRED_NAMES)}

    def _sort_key(item: AnnualFormalEmployeeInfo) -> Tuple[int, object, int, str]:
        if item.name in leader_order:
            return (0, leader_order[item.name], item.seq, item.name)
        return (1, item.seq, item.seq, item.name)

    return sorted(merged_by_name.values(), key=_sort_key)


def _parse_existing_formal_summary_rows(ws) -> Tuple[List[FormalEmployeeLeaveInfo], Dict[str, str]]:
    if ws.max_row < 4:
        return [], {}

    header_row = 3
    header_map = {
        _strip_unit_suffix(ws.cell(row=header_row, column=col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
    }
    seq_col = header_map.get("序号")
    id_col = header_map.get("工号")
    name_col = header_map.get("姓名")
    annual_total_col = header_map.get("年假总天数")
    if name_col is None:
        return [], {}

    formal_employees: List[FormalEmployeeLeaveInfo] = []
    employee_name_by_id: Dict[str, str] = {}
    fallback_seq = 1
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = _clean_text(ws.cell(row=row_idx, column=name_col).value)
        if not name:
            continue
        emp_id = _clean_text(ws.cell(row=row_idx, column=id_col).value) if id_col else ""
        seq_text = _clean_text(ws.cell(row=row_idx, column=seq_col).value) if seq_col else ""
        annual_total = _safe_float(ws.cell(row=row_idx, column=annual_total_col).value, 0.0) if annual_total_col else 0.0
        seq = int(_safe_float(seq_text, fallback_seq)) if seq_text else fallback_seq
        formal_employees.append(FormalEmployeeLeaveInfo(seq=seq, name=name, annual_leave_total=annual_total))
        if emp_id and emp_id not in employee_name_by_id:
            employee_name_by_id[emp_id] = name
        fallback_seq = max(fallback_seq + 1, seq + 1)

    return formal_employees, employee_name_by_id


def _parse_detail_sheet_employee_headers(ws) -> Dict[str, str]:
    employee_name_by_id: Dict[str, str] = {}
    if ws.max_row < 3 or ws.max_column < 4:
        return employee_name_by_id
    for col_idx in range(4, ws.max_column + 1):
        emp_id = _clean_text(ws.cell(row=2, column=col_idx).value)
        emp_name = _clean_text(ws.cell(row=3, column=col_idx).value)
        if not emp_id or not emp_name:
            continue
        employee_name_by_id[emp_id] = emp_name
    return employee_name_by_id


def _apply_half_status_to_bucket(summary: Dict[str, float], half_text: str) -> None:
    token = re.sub(r"^(上午|下午)", "", _clean_text(half_text))
    if not token:
        return
    if token == "正常":
        summary["正常数"] += 0.5
        return
    if token == "未打卡":
        summary["未打卡"] += 0.5
        return
    normalized_leave = _normalize_leave_type(token)
    if normalized_leave:
        summary[normalized_leave] += 0.5


def _apply_display_status_to_bucket(summary: Dict[str, float], display_status: object) -> None:
    status_text = _clean_text(display_status)
    if not status_text:
        return
    if status_text == "正常":
        summary["正常数"] += 1.0
        return
    if status_text == "迟到":
        summary["迟到"] += 1.0
        return
    if status_text == "早退":
        summary["早退"] += 1.0
        return
    if status_text == "迟到+早退":
        summary["迟到"] += 1.0
        summary["早退"] += 1.0
        summary["迟到+早退"] += 1.0
        return
    if status_text == "未打卡":
        summary["未打卡"] += 1.0
        return
    if status_text in {"上午未打卡", "下午未打卡"}:
        summary["未打卡"] += 0.5
        return

    normalized_leave = _normalize_leave_type(status_text)
    if normalized_leave:
        summary[normalized_leave] += 1.0
        return

    parts = [part.strip() for part in re.split(r"[，,]", status_text) if part.strip()]
    if not parts and status_text:
        parts = [status_text]
    for part in parts:
        _apply_half_status_to_bucket(summary, part)


def _rebuild_summary_by_emp_from_detail_sheet(ws) -> Dict[str, Dict[str, float]]:
    summary_by_emp: Dict[str, Dict[str, float]] = {}
    employee_ids = [
        _clean_text(ws.cell(row=2, column=col_idx).value)
        for col_idx in range(4, ws.max_column + 1)
        if _clean_text(ws.cell(row=2, column=col_idx).value)
    ]
    for emp_id in employee_ids:
        summary_by_emp[emp_id] = _new_summary_bucket()

    summary_start_row = None
    for row_idx in range(4, ws.max_row + 1):
        first_label = _clean_text(ws.cell(row=row_idx, column=1).value)
        if first_label in SUMMARY_ROWS:
            summary_start_row = row_idx
            break
    if summary_start_row is None:
        summary_start_row = ws.max_row + 1

    for col_idx in range(4, ws.max_column + 1):
        emp_id = _clean_text(ws.cell(row=2, column=col_idx).value)
        if not emp_id:
            continue
        summary = summary_by_emp.setdefault(emp_id, _new_summary_bucket())
        for row_idx in range(4, summary_start_row):
            _apply_display_status_to_bucket(summary, ws.cell(row=row_idx, column=col_idx).value)

    return summary_by_emp


def _rebuild_detail_rows_from_existing_sheet(ws) -> Tuple[List[List[object]], Dict[str, Dict[str, float]], Dict[str, str]]:
    employee_name_by_id = _parse_detail_sheet_employee_headers(ws)
    employee_ids = [
        _clean_text(ws.cell(row=2, column=col_idx).value)
        for col_idx in range(4, ws.max_column + 1)
        if _clean_text(ws.cell(row=2, column=col_idx).value)
    ]

    summary_start_row = None
    for row_idx in range(4, ws.max_row + 1):
        first_label = _clean_text(ws.cell(row=row_idx, column=1).value)
        if first_label in SUMMARY_ROWS:
            summary_start_row = row_idx
            break
    if summary_start_row is None:
        summary_start_row = ws.max_row + 1

    rows: List[List[object]] = []
    rows.append(["日期", "星期", "员工工号"] + employee_ids)
    rows.append(["", "", "员工姓名"] + [employee_name_by_id.get(emp_id, "") for emp_id in employee_ids])

    for row_idx in range(4, summary_start_row):
        row_values = [_clean_text(ws.cell(row=row_idx, column=1).value), _clean_text(ws.cell(row=row_idx, column=2).value), ""]
        for col_idx in range(4, 4 + len(employee_ids)):
            row_values.append(_clean_text(ws.cell(row=row_idx, column=col_idx).value))
        rows.append(row_values)

    summary_by_emp = _rebuild_summary_by_emp_from_detail_sheet(ws)
    for label in SUMMARY_ROWS:
        row = [label, "", ""]
        for emp_id in employee_ids:
            row.append(_format_stat_value(summary_by_emp.get(emp_id, _new_summary_bucket()).get(label, 0.0)))
        rows.append(row)

    for leave_type in LEAVE_TYPE_ROWS:
        row = [leave_type, "", ""]
        for emp_id in employee_ids:
            row.append(_format_stat_value(summary_by_emp.get(emp_id, _new_summary_bucket()).get(leave_type, 0.0)))
        rows.append(row)

    return rows, summary_by_emp, employee_name_by_id


def _replace_sheet_with_writer(wb, sheet_name: str, writer: Callable[[object], None], *, insert_index: Optional[int] = None) -> None:
    existing_index = insert_index
    if sheet_name in wb.sheetnames:
        existing_index = wb.sheetnames.index(sheet_name)
        wb.remove(wb[sheet_name])
    if existing_index is None:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb.create_sheet(sheet_name, existing_index)
    writer(ws)


def refresh_existing_result_workbook(
    result_file: str = OUTPUT_FILE,
    logger: Optional[Callable[[str], None]] = None,
) -> RefreshSummary:
    log = logger or (lambda _message: None)
    result_path = Path(result_file)
    if not result_path.exists():
        raise RuntimeError(f"未找到结果文件：{result_path}")

    log("正在读取已有结果文件...")
    wb = load_workbook(result_path)

    monthly_detail_infos: List[Tuple[int, int, str]] = []
    for sheet_name in wb.sheetnames:
        parsed = parse_monthly_detail_sheet_name(sheet_name)
        if parsed is None:
            continue
        monthly_detail_infos.append((parsed[0], parsed[1], sheet_name))
    monthly_detail_infos.sort()
    if not monthly_detail_infos:
        raise RuntimeError("结果文件中未找到任何“YYYY-MM考勤明细”sheet。")

    grouped_by_year: Dict[int, List[Tuple[int, int, str]]] = defaultdict(list)
    for year, month, sheet_name in monthly_detail_infos:
        grouped_by_year[year].append((year, month, sheet_name))

    refreshed_months: List[Tuple[int, int]] = []
    refreshed_years: List[int] = []

    for year in sorted(grouped_by_year):
        month_entries = grouped_by_year[year]
        log(f"正在刷新 {year} 年汇总...")
        merged_employee_name_by_id: Dict[str, str] = {}
        annual_summary_by_emp: Dict[str, Dict[str, float]] = defaultdict(_new_summary_bucket)
        monthly_results: List[MonthlyProcessedResult] = []

        annual_sheet_name = get_annual_sheet_name(year)
        annual_formal_fallback: List[FormalEmployeeLeaveInfo] = []
        if annual_sheet_name in wb.sheetnames:
            annual_formal_fallback, _ = _parse_existing_formal_summary_rows(wb[annual_sheet_name])

        for index, (_, month, detail_sheet_name) in enumerate(month_entries, start=1):
            log(f"[{index}/{len(month_entries)}] 正在刷新 {year}-{month:02d} 月度汇总...")
            detail_ws = wb[detail_sheet_name]
            detail_rows, summary_by_emp, detail_employee_name_by_id = _rebuild_detail_rows_from_existing_sheet(detail_ws)
            merged_employee_name_by_id.update(detail_employee_name_by_id)

            _replace_sheet_with_writer(
                wb,
                detail_sheet_name,
                lambda ws, rows=detail_rows, title=MONTHLY_DETAIL_TITLE_TEMPLATE.format(year=year, month=month): _write_detail_sheet(ws, rows, title),
            )

            summary_sheet_name = get_monthly_summary_sheet_name(year, month)
            monthly_formal_employees: List[FormalEmployeeLeaveInfo] = []
            summary_employee_name_by_id: Dict[str, str] = {}
            if summary_sheet_name in wb.sheetnames:
                monthly_formal_employees, summary_employee_name_by_id = _parse_existing_formal_summary_rows(wb[summary_sheet_name])
                merged_employee_name_by_id.update(summary_employee_name_by_id)
            if not monthly_formal_employees:
                monthly_formal_employees = list(annual_formal_fallback)
            if not monthly_formal_employees:
                fallback_seq = 1
                for emp_id, emp_name in sorted(detail_employee_name_by_id.items(), key=lambda item: (int(item[0]) if item[0].isdigit() else item[0])):
                    monthly_formal_employees.append(FormalEmployeeLeaveInfo(seq=fallback_seq, name=emp_name, annual_leave_total=0.0))
                    fallback_seq += 1

            monthly_summary_rows = build_monthly_formal_summary_rows(
                monthly_formal_employees,
                merged_employee_name_by_id,
                summary_by_emp,
            )
            _replace_sheet_with_writer(
                wb,
                summary_sheet_name,
                lambda ws, rows=monthly_summary_rows, title=FORMAL_TITLE_TEMPLATE.format(year=year, month=month): _write_formal_summary_sheet(ws, rows, title),
            )

            bundle = MonthlySourceBundle(
                year=year,
                month=month,
                attendance_file=result_path,
                leave_file=result_path,
                annual_leave_file=result_path,
            )
            monthly_results.append(
                MonthlyProcessedResult(
                    bundle=bundle,
                    employee_name_by_id=dict(detail_employee_name_by_id),
                    formal_employees=monthly_formal_employees,
                    report_rows=[],
                    formal_summary_rows=monthly_summary_rows,
                    summary_by_emp=summary_by_emp,
                    workday_source="refresh_from_result",
                    leave_record_count=0,
                )
            )
            for emp_id, summary in summary_by_emp.items():
                for key, value in summary.items():
                    annual_summary_by_emp[emp_id][key] += float(value)
            refreshed_months.append((year, month))
            log(f"[{index}/{len(month_entries)}] 已刷新 {year}-{month:02d} 月度汇总")

        annual_formal_employees = merge_annual_formal_employees(monthly_results)
        annual_rows = build_annual_summary_rows(
            annual_formal_employees,
            merged_employee_name_by_id,
            dict(annual_summary_by_emp),
        )
        _replace_sheet_with_writer(
            wb,
            annual_sheet_name,
            lambda ws, rows=annual_rows, title=ANNUAL_TITLE_TEMPLATE.format(year=year): _write_formal_summary_sheet(ws, rows, title),
            insert_index=0,
        )
        refreshed_years.append(year)
        log(f"已刷新 {year} 年度汇总")

    log("正在保存刷新后的结果文件...")
    wb.save(result_path)
    log(f"已刷新: {result_path}")
    return RefreshSummary(
        years=refreshed_years,
        refreshed_months=refreshed_months,
        output_file=result_path,
    )


def build_report(
    employee_name_by_id: Dict[str, str],
    records: Dict[str, Dict[date, Dict[str, object]]],
    year: int,
    month: int,
    year_workday_set: set[date],
) -> Tuple[List[List[object]], Dict[str, Dict[str, float]]]:
    """构造最终月度汇总表（日期明细 + 汇总统计）。"""
    _, days_in_month = calendar.monthrange(year, month)
    all_days = [date(year, month, day) for day in range(1, days_in_month + 1)]
    workday_set = {d for d in all_days if is_workday(d, year_workday_set)}

    employee_ids = sorted(employee_name_by_id.keys(), key=lambda x: (int(x) if x.isdigit() else x))

    rows: List[List[object]] = []
    rows.append(["日期", "星期", "员工工号"] + employee_ids)
    rows.append(["", "", "员工姓名"] + [employee_name_by_id.get(emp_id, "") for emp_id in employee_ids])

    for current_date in all_days:
        row = [current_date.isoformat(), WEEKDAY_LABELS[current_date.weekday()], ""]
        is_current_workday = current_date in workday_set
        for emp_id in employee_ids:
            if not is_current_workday:
                row.append("")
                continue
            day_record = records.get(emp_id, {}).get(current_date)
            row.append(day_record.get("display_status", day_record["status"]) if day_record else MISSING_DAY_OUTPUT)
        rows.append(row)

    summary_by_emp = build_summary_by_employee(employee_ids, records, all_days, workday_set)

    for label in SUMMARY_ROWS:
        row = [label, "", ""]
        for emp_id in employee_ids:
            row.append(_format_stat_value(summary_by_emp[emp_id][label]))
        rows.append(row)

    for leave_type in LEAVE_TYPE_ROWS:
        row = [leave_type, "", ""]
        for emp_id in employee_ids:
            row.append(_format_stat_value(summary_by_emp[emp_id][leave_type]))
        rows.append(row)

    return rows, summary_by_emp


def process_monthly_bundle(
    bundle: MonthlySourceBundle,
) -> MonthlyProcessedResult:
    sheet_map = read_input_file(str(bundle.attendance_file))
    blocks = normalize_columns(sheet_map)
    if not blocks:
        raise RuntimeError(f"未识别到员工数据块，请检查文件: {bundle.attendance_file}")

    year_workday_set, workday_source = get_year_workday_set(bundle.year)
    employee_name_by_id, records = parse_attendance_records(sheet_map, blocks, bundle.year, bundle.month)
    leave_df = read_leave_file(str(bundle.leave_file))
    annual_leave_df = read_annual_leave_file(str(bundle.annual_leave_file))
    formal_employees = parse_formal_employee_leave_info(annual_leave_df)
    leave_map = parse_leave_records(leave_df, employee_name_by_id, bundle.year, bundle.month)
    merged_records = apply_leave_to_attendance(records, leave_map)
    report_rows, summary_by_emp = build_report(
        employee_name_by_id,
        merged_records,
        bundle.year,
        bundle.month,
        year_workday_set,
    )
    formal_summary_rows = build_monthly_formal_summary_rows(formal_employees, employee_name_by_id, summary_by_emp)
    leave_record_count = sum(len(v) for v in leave_map.values())
    return MonthlyProcessedResult(
        bundle=bundle,
        employee_name_by_id=employee_name_by_id,
        formal_employees=formal_employees,
        report_rows=report_rows,
        formal_summary_rows=formal_summary_rows,
        summary_by_emp=summary_by_emp,
        workday_source=workday_source,
        leave_record_count=leave_record_count,
    )


def _write_detail_sheet(ws, rows: List[List[object]], title_text: str) -> None:
    max_col = len(rows[0]) if rows else 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value=title_text)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    max_col = ws.max_column
    max_row = ws.max_row

    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")

    summary_start_row = max_row - (len(SUMMARY_ROWS) + len(LEAVE_TYPE_ROWS)) + 1
    for r_idx in range(summary_start_row, max_row + 1):
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=3)

    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    title_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    for r_idx in range(1, max_row + 1):
        for c_idx in range(1, max_col + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = center
            if r_idx == 1:
                cell.alignment = title_alignment
                cell.font = title_font
            elif r_idx <= 3 or c_idx == 1:
                cell.font = bold

    for r_idx in range(1, max_row + 1):
        row_label = _clean_text(ws.cell(row=r_idx, column=1).value)
        row_fill_color = (
            MISSING_HIGHLIGHT_COLOR
            if row_label == "未打卡"
            else STATUS_FILL_COLORS.get(row_label)
        )

        if row_fill_color:
            header_cols = (1, 2, 3) if r_idx >= summary_start_row else (1, 2)
            for c_idx in header_cols:
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill(
                    fill_type="solid",
                    fgColor=row_fill_color,
                )

        for c_idx in range(1, max_col + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            value_text = _clean_text(cell.value)
            detail_fill_key = _extract_fill_key(value_text) if r_idx >= 4 and c_idx >= 3 else None

            if detail_fill_key == "未打卡":
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=MISSING_HIGHLIGHT_COLOR,
                )
            elif detail_fill_key and detail_fill_key in STATUS_FILL_COLORS:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=STATUS_FILL_COLORS[detail_fill_key],
                )
            elif row_fill_color and c_idx >= 3 and _is_nonzero_number_text(value_text):
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=row_fill_color,
                )

    ws.freeze_panes = "D4"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    for c_idx in range(4, max_col + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 10


def _write_formal_summary_sheet(
    ws,
    rows: List[List[object]],
    title_text: str,
) -> None:
    note = "注：正常数、未打卡及各类假期单位为天；迟到、早退、迟到+早退单位为次。"
    max_col = len(rows[0]) if rows else 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=1, column=1, value=title_text)
    ws.cell(row=2, column=1, value=note)

    display_rows: List[List[object]] = []
    for idx, row in enumerate(rows):
        if idx == 0:
            display_rows.append([_wrap_header_text(value) for value in row])
        else:
            display_rows.append(row)

    for r_idx, row in enumerate(display_rows, start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)

    header_row = 3
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[header_row].height = 32
    header_fill_key_map = {
        "正常数": None,
        "状态": None,
        "迟到": "迟到",
        "早退": "早退",
        "迟到+早退": "迟到+早退",
        "年假": "年假",
        "年假（奖励）": "年假（奖励）",
        "年假（其他）": "年假（其他）",
        "事假": "事假",
        "病假": "病假",
        "未打卡": "未打卡",
        "事+病+未打卡总计": None,
        "婚假": "婚假",
        "产假": "产假",
        "丧假": "丧假",
        "探亲假": "探亲假",
        "公假": "公假",
        "育儿假": "育儿假",
        "年假总天数": None,
        "结余年假总天数": None,
        FORMAL_NON_PUBLIC_LEAVE_TOTAL_LABEL: None,
    }
    header_group_map = {
        "序号": "identity",
        "工号": "identity",
        "姓名": "identity",
        "状态": "identity",
        "正常数": "normal",
        "迟到": "attendance",
        "早退": "attendance",
        "迟到+早退": "attendance",
        "事假": "absence",
        "病假": "absence",
        "未打卡": "absence",
        "事+病+未打卡总计": "absence",
        "婚假": "leave",
        "产假": "leave",
        "丧假": "leave",
        "探亲假": "leave",
        "公假": "leave",
        "育儿假": "leave",
        "年假": "annual",
        "年假（奖励）": "annual",
        "年假（其他）": "annual",
        "年假总天数": "annual",
        "结余年假总天数": "annual",
        FORMAL_NON_PUBLIC_LEAVE_TOTAL_LABEL: "leave",
    }
    header_map = {idx: _clean_text(rows[0][idx - 1]) for idx in range(1, max_col + 1)}

    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, max_col + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if r_idx != 2:
                cell.alignment = center
            if r_idx == header_row:
                cell.font = bold
                header_label = header_map.get(c_idx, "")
                header_group = header_group_map.get(header_label, "identity")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=FORMAL_HEADER_FILL_COLORS[header_group],
                )

            if r_idx > header_row:
                col_label = header_map.get(c_idx, "")
                value_text = _clean_text(cell.value)
                fill_key = header_fill_key_map.get(col_label)
                if fill_key and _is_nonzero_number_text(value_text):
                    fill_color = (
                        MISSING_HIGHLIGHT_COLOR
                        if fill_key == "未打卡"
                        else STATUS_FILL_COLORS.get(fill_key)
                    )
                    if fill_color:
                        cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

    width_map = {
        "A": 8,
        "B": 12,
        "C": 12,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width
    for c_idx in range(4, max_col + 1):
        header = header_map.get(c_idx, "")
        ws.column_dimensions[get_column_letter(c_idx)].width = 14 if header in {"迟到+早退", "事+病+未打卡总计", "结余年假总天数", FORMAL_NON_PUBLIC_LEAVE_TOTAL_LABEL} else 10

    ws.freeze_panes = "A4"


def export_report(
    monthly_results: List[MonthlyProcessedResult],
    annual_rows: List[List[object]],
    year: int,
    output_file: str = OUTPUT_FILE,
) -> None:
    """导出考勤统计结果.xlsx。"""
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    annual_sheet_name = get_annual_sheet_name(year)
    annual_ws = wb.create_sheet(annual_sheet_name)
    annual_title = ANNUAL_TITLE_TEMPLATE.format(year=year)
    _write_formal_summary_sheet(annual_ws, annual_rows, annual_title)

    for monthly_result in sorted(monthly_results, key=lambda item: (item.bundle.year, item.bundle.month)):
        month = monthly_result.bundle.month
        monthly_detail_sheet_name = get_monthly_detail_sheet_name(year, month)
        monthly_summary_sheet_name = get_monthly_summary_sheet_name(year, month)

        detail_title = MONTHLY_DETAIL_TITLE_TEMPLATE.format(year=year, month=month)
        detail_ws = wb.create_sheet(monthly_detail_sheet_name)
        _write_detail_sheet(detail_ws, monthly_result.report_rows, detail_title)

        monthly_title = FORMAL_TITLE_TEMPLATE.format(year=year, month=month)
        summary_ws = wb.create_sheet(monthly_summary_sheet_name)
        _write_formal_summary_sheet(summary_ws, monthly_result.formal_summary_rows, monthly_title)

    wb.save(output_file)


def generate_report(
    data_dir: str = DATA_DIR,
    output_file: str = OUTPUT_FILE,
    logger: Optional[Callable[[str], None]] = None,
    target_year: Optional[int] = None,
    relaxed: bool = False,
) -> ReportSummary:
    log = logger or (lambda _message: None)

    log("正在扫描可统计月份...")
    bundles = discover_monthly_source_bundles(data_dir, target_year=target_year, relaxed=relaxed)
    if not bundles:
        raise RuntimeError("未发现可统计的月度文件。")

    year = bundles[0].year
    log(f"已识别 {len(bundles)} 个月份，开始逐月汇总...")
    monthly_results: List[MonthlyProcessedResult] = []
    merged_employee_name_by_id: Dict[str, str] = {}
    annual_summary_by_emp: Dict[str, Dict[str, float]] = defaultdict(_new_summary_bucket)
    formal_employees: List[FormalEmployeeLeaveInfo] = []

    for index, bundle in enumerate(bundles, start=1):
        log(f"[{index}/{len(bundles)}] 正在处理 {bundle.year}-{bundle.month:02d}...")
        monthly_result = process_monthly_bundle(bundle)
        monthly_results.append(monthly_result)
        merged_employee_name_by_id.update(monthly_result.employee_name_by_id)
        formal_employees = monthly_result.formal_employees
        for emp_id, summary in monthly_result.summary_by_emp.items():
            for key, value in summary.items():
                annual_summary_by_emp[emp_id][key] += float(value)
        log(f"[{index}/{len(bundles)}] 已完成 {bundle.year}-{bundle.month:02d}")

    log("正在汇总年度正式员工数据...")
    annual_formal_employees = merge_annual_formal_employees(monthly_results)
    annual_rows = build_annual_summary_rows(
        annual_formal_employees,
        merged_employee_name_by_id,
        dict(annual_summary_by_emp),
    )
    log("正在写入 Excel 文件...")
    export_report(monthly_results, annual_rows, year, output_file)
    log("正在整理导出结果...")

    log(f"已生成: {output_file}")
    log(f"统计年份: {year}")
    log(f"统计月份数: {len(monthly_results)}")
    log(f"正式员工数(含领导): {len(annual_formal_employees)}")
    for monthly_result in monthly_results:
        bundle = monthly_result.bundle
        year_workday_set, _ = get_year_workday_set(bundle.year)
        month_workday_count = sum(1 for d in year_workday_set if d.year == bundle.year and d.month == bundle.month)
        log(
            f"{bundle.year}-{bundle.month:02d}: "
            f"员工数={len(monthly_result.employee_name_by_id)}, "
            f"工作日来源={monthly_result.workday_source}, "
            f"工作日数量={month_workday_count}, "
            f"请假记录数={monthly_result.leave_record_count}"
        )

    return ReportSummary(
        year=year,
        monthly_results=monthly_results,
        formal_employee_count=len(annual_formal_employees),
        output_file=Path(output_file),
    )


def main() -> None:
    generate_report(DATA_DIR, OUTPUT_FILE, print)


if __name__ == "__main__":
    main()
